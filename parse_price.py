import openpyxl
import json

wb = openpyxl.load_workbook('d:/dairy_first_project/price.xlsx', data_only=True)
sheet = wb.active

data = []
for r in range(1, sheet.max_row + 1):
    row_vals = []
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=r, column=c).value
        row_vals.append(v)
    data.append(row_vals)

with open('d:/dairy_first_project/price_parsed.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Parsed {len(data)} rows successfully!")
