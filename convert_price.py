import openpyxl
import json
import re

marathi_to_english_map = {
    '०': '0', '१': '1', '२': '2', '३': '3', '४': '4',
    '५': '5', '६': '6', '७': '7', '८': '8', '९': '9'
}

def parse_num(val):
    if val is None:
        return None
    val_str = str(val).strip()
    for m, e in marathi_to_english_map.items():
        val_str = val_str.replace(m, e)
    try:
        return float(val_str)
    except ValueError:
        return val_str

wb = openpyxl.load_workbook('d:/dairy_first_project/price.xlsx', data_only=True)
sheet = wb.active

parsed_chart = []

# Row 1 has SNF headers
header_row = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
snf_headers = []
for col_idx, raw_snf in enumerate(header_row[1:], start=2):
    parsed_snf = parse_num(raw_snf)
    if isinstance(parsed_snf, float):
        snf_headers.append((col_idx, parsed_snf))

print("Detected SNF Columns:", [snf for _, snf in snf_headers])

rate_entries = []

for r in range(2, sheet.max_row + 1):
    raw_fat = sheet.cell(row=r, column=1).value
    fat_val = parse_num(raw_fat)
    if not isinstance(fat_val, float):
        continue
    
    for col_idx, snf_val in snf_headers:
        raw_rate = sheet.cell(row=r, column=col_idx).value
        rate_val = parse_num(raw_rate)
        if isinstance(rate_val, float) and rate_val > 0:
            rate_entries.append({
                'fat': fat_val,
                'snf': snf_val,
                'rate': rate_val
            })

print(f"Total valid rate entries parsed: {len(rate_entries)}")
if rate_entries:
    fats = sorted(list(set(e['fat'] for e in rate_entries)))
    snfs = sorted(list(set(e['snf'] for e in rate_entries)))
    print(f"FAT Range: min={min(fats)}, max={max(fats)}, count={len(fats)}")
    print(f"SNF Range: min={min(snfs)}, max={max(snfs)}, count={len(snfs)}")
    print("Sample Entries:", rate_entries[:5])

with open('d:/dairy_first_project/rate_entries.json', 'w', encoding='utf-8') as f:
    json.dump(rate_entries, f, indent=2)
